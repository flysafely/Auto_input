from pymouse import PyMouse
from pykeyboard import PyKeyboard
from pykeyboard import PyKeyboardEvent

from openpyxl import Workbook  # 写excel时候用，因为win32com使用多线程的时候会提示 被呼叫方拒绝接收呼叫 的错误
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

from win32com.client import GetObject
from win32com.client import Dispatch
from win32com.client import constants as constants
from oscrypto._win import symmetric
from PIL import Image, ImageTk

from tkinter import *
from tkinter import ttk
from tkinter.filedialog import *
from tkinter.messagebox import *

import win32gui
import win32api
import win32con
import win32process
import tkinter
import time
import ctypes
import inspect
import win32com.client
import threading
import pythoncom  # 多线程调用COM
import platform
import os
import time
import datetime
import uuid
import urllib.request
import re
import CheckRegister as ckr
import CheckUpdate as cku

Version = "2.1"
Software_Name = "ae"

m = PyMouse()
k = PyKeyboard()

global time_stamp, isOk, isContinue, isStop, isRegistered, Status_label, data_array, hllDll, VK_CAPITAL, VK_NUMLOCK, VK_SCROLL, VK_ESCAPE, UserName, Company, Department
isContinue = True
isStop = False
isOk = False
isRegistered = False
isOver = False
time_stamp = 0
hllDll = ctypes.WinDLL("User32.dll")
key_hex = [0x1B, 0x23, 0x14, 0x90, 0x91]
VK_ESCAPE = key_hex[0]
VK_END = key_hex[1]
VK_CAPITAL = key_hex[2]
VK_NUMLOCK = key_hex[3]
VK_SCROLL = key_hex[4]

ADD_REG_EXP = "now : '(.*?) "  # 获取内网OA中的时间戳信息
server_ip = "http://130.130.200.49"
column_var = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
              "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT"]
title_var = ["专柜", "合同号", "时间"]
pay_num = ["B100", "B102", "B110", "B112", "B120", "B122", "B130", "B132", "B200", "B202",
           "C100", "C102", "C110", "C112", "C120", "C122", "C130", "C132", "C140", "C142", "C150", "C152", "C200", "C202", "C210", "C212", "C220", "C222", "C230", "C232",
           "C240", "C242", "C250", "C252", "C260", "C262", "C270", "C272", "C280", "C282", "C290", "C292", "C900",
           "D010", "D020",
           "E100", "E900",
           "F100", "F102", "F110", "F112", "F120",
           "V100", "V110", "V120", "V210", "V220", "V230", "V250", "V260", "V270", "V280", "V300"]

pay_mode = ["Z", "X", "C", "z", "x", "c"]
data_array = []


class myThread (threading.Thread):

    def __init__(self, functions):
        threading.Thread.__init__(self)
        self.functions = functions
        self.result = object

    def run(self):
        self.functions()

    def get_result(self):
        return self.result


def Add_Thread(function):
    thread = myThread(function)
    thread.setDaemon(True)
    thread.start()
    return thread


def operation_with_pause(mode, *args, **kwargs):

    if mode == "pk":
        for i in range(1, kwargs["num"] + 1):
            k.tap_key(args[0])

    elif mode == "ps":
        for i in range(1, kwargs["num"] + 1):
            k.type_string(args[0])

    elif mode == "pks":
        for i in range(1, kwargs["num"] + 1):
            k.press_keys(list(args))


def setAppWindowForeground(Appname, windowsize=3):  # 让特定程序获得焦点
    hForeWnd = win32gui.GetForegroundWindow()
    dwCurID = win32api.GetCurrentThreadId()
    dwForeID = win32process.GetWindowThreadProcessId(hForeWnd)

    windowHandle = win32gui.FindWindow(0, Appname)
    if hForeWnd != 0 and dwCurID != dwForeID[0]:
        win32process.AttachThreadInput(dwCurID, dwForeID[0], True)
    # 参数解释https://msdn.microsoft.com/en-us/library/windows/desktop/ms633548(v=vs.85).aspx
    win32gui.ShowWindow(windowHandle, windowsize)
    win32gui.SetWindowPos(windowHandle, -1, 0, 0, 0, 0, 0x0001 | 0x0002)
    win32gui.SetWindowPos(windowHandle, -2, 0, 0, 0, 0, 0x0001 | 0x0002)
    if hForeWnd != 0:
        win32gui.SetForegroundWindow(windowHandle)


def ensure_App_Foreground(appname, windowsize=3):
    if win32gui.FindWindow(0, appname) != win32gui.GetForegroundWindow():
        setAppWindowForeground(appname, windowsize)


def ensure_CapsLock():
    global hllDll, VK_CAPITAL, isContinue
    if hllDll.GetKeyState(VK_CAPITAL) == 0:
        k.tap_key(k.caps_lock_key)


def check_window(appname):
    global isStop
    if win32gui.GetForegroundWindow() != win32gui.FindWindow(0, appname):
        isStop = True


def keyboard_monitor(hexnum, function):
    global hllDll, isContinue
    if hexnum in key_hex and hexnum != 0x23:
        pre_status = hllDll.GetKeyState(hexnum)
        while 1:
            after_status = hllDll.GetKeyState(hexnum)

            # 0,1状态为up，65408,65409为dowm状态
            if pre_status != after_status and (after_status == 65408 or after_status == 65409):
                function()  # 使用lambda表达式来传递每个按键执行的功能
                pre_status = hllDll.GetKeyState(hexnum)
    else:
        pre_status = hllDll.GetKeyState(hexnum)
        while 1:
            after_status = hllDll.GetKeyState(hexnum)

            # 0,1状态为up，-127,-128为dowm状态  end键
            if pre_status != after_status and (after_status == -128 or after_status == -127):
                function()  # 使用lambda表达式来传递每个按键执行的功能
                pre_status = hllDll.GetKeyState(hexnum)


def Change_isContinue():
    global isContinue, time_stamp
    if time.time() - time_stamp > 0.05:
        time_stamp = time.time()
        isContinue = not isContinue
    else:
        pass


def Change_isStop():
    global isStop
    isStop = True


def Open_WorkBook_By_Openpyxl(path):
    try:
        Workbook = load_workbook(path)
        return Workbook

    except Exception as e:
        tkinter.messagebox.showinfo(
            "表格数据有误！", "请检查导出文件%s格式、内容是否正确，亦或是没有选择对应的正确文件！" % path)
        raise


def Open_Sheet_By_Openpyxl(path, offset):  # openpyxl sheet序号是从0开始的
    try:
        Workbook = load_workbook(path)
        return (Workbook, Workbook.worksheets[0 + offset])

    except Exception as e:
        tkinter.messagebox.showinfo(
            "表格数据有误！", "请检查导出文件%s格式、内容是否正确，亦或是没有选择对应的正确文件！" % path)
        raise


def Get_Max_Range_Num(worksheet, method):
    if method == "row":
        num = 0
        for i in range(1, worksheet.max_row + 1):
            if worksheet.cell(row=i, column=1).value != None:
                num += 1
        return num
    else:
        num = 0
        for i in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=1, column=i).value != None:
                num += 1
        return num


def is_num_by_except(num, message, v_num):
    try:
        float(num)
        return True
    except Exception as e:
        v_num.set("警告！", message + "单元格内容为非数字!")
        tkinter.messagebox.showinfo("警告！", message + "单元格内容为非数字!")
        return False


def get_path(label):
    label.set("正在处理导入数据......")
    choose_path = tkinter.filedialog.askopenfilename()
    if choose_path:
        choose_path = choose_path.replace("/", "\\")
        check_data(choose_path, 0, label)
    else:
        label.set("等待导入数据......")


def check_data(path, offset, label):
    global isOk
    isOk = True
    Excel_WorkBook, Excel_WorkSheet = Open_Sheet_By_Openpyxl(path, offset)

    Max_Column = Get_Max_Range_Num(Excel_WorkSheet, "column")
    Max_Row = Get_Max_Range_Num(Excel_WorkSheet, "row")
    # 检查表格中是否有数据
    if Max_Row < 2 or Max_Column < 4:
        tkinter.messagebox.showinfo("警告！", "未检测到导入表格中的有效数据！")
        return
    # 检查是否存在有数据，但是没有字段名称的情况
    for column_num in range(1, Max_Column + 1):
        Cell_Value = Excel_WorkSheet.cell(row=1, column=column_num).value
        if Cell_Value == None:
            tkinter.messagebox.showinfo("警告！", "%s列:'%s'<标题>为空!" % (
                column_var[column_num - 1], Cell_Value))
            isOk = False
            return
        else:
            if column_num < 4:
                if Cell_Value != title_var[column_num - 1]:
                    tkinter.messagebox.showinfo("警告！", "%s列 : '%s' <标题>错误!,正确为: '%s' " % (
                        column_var[column_num - 1], Cell_Value, title_var[column_num - 1]))
                    label.set("%s列 : '%s' <标题>错误!,正确为: '%s' " % (
                        column_var[column_num - 1], Cell_Value, title_var[column_num - 1]))
                    isOk = False
                    return
            elif Cell_Value[0:4] not in pay_num:
                tkinter.messagebox.showinfo("警告！", "%s列:'%s'<费用编码>错误!" % (
                    column_var[column_num - 1], Cell_Value))
                label.set("%s列:'%s'<费用编码>错误!" %
                          (column_var[column_num - 1], Cell_Value))
                isOk = False
                return

    for row_num in range(2, Max_Row + 1):
        Cell_Value1 = Excel_WorkSheet.cell(row=row_num, column=2).value
        Cell_Value2 = Excel_WorkSheet.cell(row=row_num, column=3).value
        if Cell_Value1 != None and str(Cell_Value1).isdigit() and Cell_Value2 != None:
            date_detail = str(Cell_Value2).split(" ")[0].split("-")
            if len(date_detail) == 3:
                if len(date_detail[0]) + len(date_detail[1]) + len(date_detail[2]) != 8:
                    tkinter.messagebox.showinfo(
                        "警告！", "C%d单元格日期格式有误！例如:2017-08-25" % row_num)
                    label.set("C%d单元格日期格式有误！例如:2017-08-25" % row_num)
                    isOk = False
                    return
            else:
                tkinter.messagebox.showinfo(
                    "警告！", "C%d单元格日期分割符号有误！分隔符为'-'" % row_num)
                label.set("C%d单元格日期分割符号有误！分隔符为'-'" % row_num)
                isOk = False
                return
        else:
            tkinter.messagebox.showinfo(
                "警告！", "%s行:<合同号>或者<录入时间>错误!" % str(row_num))
            label.set("%s行:<合同号>或者<录入时间>错误!" % str(row_num))
            isOk = False
            return

    get_excel_data(Excel_WorkSheet, Max_Column, Max_Row, label)


def get_excel_data(Excel_WorkSheet, Max_Column, Max_Row, label):
    global data_array

    title_array = []
    contract_error = []
    value_error_row = ""
    con_error_row = ""
    all_data_array = []

    for i in range(2, Max_Column + 1):
        title_array.append(Excel_WorkSheet.cell(row=1, column=i).value)
    for r in range(2, Max_Row + 1):
        row_array = []
        for j in range(2,  Max_Column + 1):
            if j <= 3:
                row_array.append((title_array[j - 2], str(Excel_WorkSheet.cell(
                    row=r, column=j).value).replace(".0", "")))

            elif j > 3 and j % 2 == 0:
                cell_value = str(Excel_WorkSheet.cell(row=r, column=j).value)
                cell_value_without_char = cell_value.replace(
                    ".", "", 1).replace("-", "", 1)
                if cell_value not in ["None", "0", "0.0"] and cell_value_without_char.isdigit() and cell_value_without_char != "0":
                    if float(Excel_WorkSheet.cell(row=r, column=j).value) > 0:
                        if str(Excel_WorkSheet.cell(row=r, column=j + 1).value) in pay_mode:
                            row_array.append((title_array[j - 2],
                                              str(Excel_WorkSheet.cell(
                                                  row=r, column=j).value),
                                              str(Excel_WorkSheet.cell(row=r, column=j + 1).value).upper()))
                        else:
                            row_array.append((title_array[j - 2],
                                              str(Excel_WorkSheet.cell(
                                                  row=r, column=j).value),
                                              "Z"))
                    else:
                        if str(Excel_WorkSheet.cell(row=r, column=j + 1).value) in pay_mode:
                            row_array.append((title_array[j - 2][:3] + "2",
                                              str(abs(float(Excel_WorkSheet.cell(
                                                  row=r, column=j).value))),
                                              str(Excel_WorkSheet.cell(row=r, column=j + 1).value).upper()))
                        else:
                            row_array.append((title_array[j - 2][:3] + "2",
                                              str(abs(float(Excel_WorkSheet.cell(
                                                  row=r, column=j).value))),
                                              "Z"))

        all_data_array.append(row_array)

    data_array = list(filter(lambda data: len(data) > 2, all_data_array))
    print(isOk)
    label.set("导入数据检查通过！")


def start_input(second, pay_windows, rowinfo, pending):
    global isOk, isStop, isRegistered, hllDll, VK_CAPITAL, VK_NUMLOCK, VK_SCROLL, VK_ESCAPE, data_array
    if isRegistered == False:
        pay_windows.deiconify()
        if tkinter.messagebox.askyesno("警告！", "软件'未激活'或者'注册码已过期',请扫描左侧二维码！"):
            pass
        else:
            pay_windows.withdraw()
    else:
        if len(data_array) != 0 and len(second.split("-")) == 2 and second.split("-")[0].isdigit() and second.split("-")[1].replace("0.", "").isdigit() and isOk:
            wait_time = second.split("-")[0]
            trim_time = second.split("-")[1]
            if tkinter.messagebox.askyesno("提示！", "即将开始录入数据！请确认录入窗口已经打开！"):
                try:
                    ensure_App_Foreground(u"富基融通商业连锁门店管理系统", 3)
                except Exception as e:
                    tkinter.messagebox.showinfo("警告！", "尚未打开费用录入单界面！")
                else:
                    if analyze_rowinfo(rowinfo)[0] == -1:
                        tkinter.messagebox.showinfo(
                            "警告！", analyze_rowinfo(rowinfo)[1])
                    else:
                        time.sleep(1)
                        check_window(u"富基融通商业连锁门店管理系统")
                        ensure_CapsLock()
                        program = Simulation_operation(
                            int(wait_time), analyze_rowinfo(rowinfo), pending)
                        do_and_check_pause(
                            program, float(trim_time), True, False)
            else:
                print("错误！")
        else:
            tkinter.messagebox.showinfo("警告！", "<延迟秒速>或<起始行>或<费用数据>有误,请检查！")


def do_and_check_pause(program, trim, iscontinue, isstop):
    global isContinue, isStop
    isContinue = iscontinue
    isStop = isstop
    wait_time = time.time()
    while 1:
        if isContinue == True and isStop == False:
            try:
                next(program)
                time.sleep(trim)
            except StopIteration:
                tkinter.messagebox.showinfo("提示！", "录入已完成！")
                break
        elif isContinue == True and isStop == True:
            break
        elif isContinue == False and isStop == True:
            break
        else:
            if time.time() - wait_time <= 60:
                pass
            else:
                print("wait out!")
                break
    tkinter.messagebox.showinfo("警告！", "已经终止录入！")


def analyze_rowinfo(rowinfo):
    global data_array
    if "*" in rowinfo:
        if int(rowinfo.replace("*", "")) >= 2 and int(rowinfo.replace("*", "")) <= len(data_array) + 1:
            return [int(rowinfo.replace("*", "")) - 2, int(rowinfo.replace("*", "")) - 1]
        else:
            return [-1, "单行录入超出<起始>或<结束>位置！"]

    elif "-" in rowinfo:
        if int(rowinfo.split("-")[0]) >= 2 and int(rowinfo.split("-")[0]) <= int(rowinfo.split("-")[1]) and int(rowinfo.split("-")[1]) <= len(data_array) + 1:
            return [int(rowinfo.split("-")[0]) - 2, int(rowinfo.split("-")[1]) - 1]
        elif int(rowinfo.split("-")[0]) < 2 and int(rowinfo.split("-")[1]) <= len(data_array) + 1:
            return [-1, "指定行录入<起始>位置超出范围！"]
        elif int(rowinfo.split("-")[0]) > 2 and int(rowinfo.split("-")[1]) > len(data_array) + 1:
            return [-1, "指定行录入<结束>位置超出范围！"]
    elif rowinfo.isdigit():
        if int(rowinfo) >= 2 and int(rowinfo) <= len(data_array) + 1:
            return [int(rowinfo) - 2, len(data_array)]
        else:
            return [-1, "指定行录入<起始>位置超出范围！"]
    else:
        return [-1, "<起始行>信息输入有误！"]


def Add_thread(function):
    thread = myThread(function)
    thread.setDaemon(True)
    thread.start()
    return thread


def _async_raise(tid, exctype):  # 用于退出子线程
    """raises the exception, performs cleanup if needed"""
    tid = ctypes.c_long(tid)
    if not inspect.isclass(exctype):
        exctype = type(exctype)
    res = ctypes.pythonapi.PyThreadState_SetAsyncExc(
        tid, ctypes.py_object(exctype))
    if res == 0:
        raise ValueError("invalid thread id")
    elif res != 1:
        # """if it returns a number greater than one, you're in trouble,
        # and you should call it again with exc=NULL to revert the effect"""
        ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, None)
        raise SystemError("PyThreadState_SetAsyncExc failed")


def Stop_thread(thread):  # 用于退出子线程
    _async_raise(thread.ident, SystemExit)


def Simulation_operation(stoptime, rowinfo, pending):
    global data_array
    # 开启检测capslocks(暂停功能)状态
    keyboard_monitor_thread = Add_thread(
        lambda: keyboard_monitor(VK_CAPITAL, lambda: Change_isContinue()))
    # 开启检测END(停止功能)状态
    os_exit_byEnd_thread = Add_thread(
        lambda: keyboard_monitor(VK_ESCAPE, lambda: Change_isStop()))
    # 开启检测ESC(退出功能)状态
    os_exit_thread = Add_thread(
        lambda: keyboard_monitor(VK_END, lambda: os._exit(0)))
    # 获取屏幕尺寸
    x_dim, y_dim = m.screen_size()
    # 当富基全屏后，点击屏幕中间位置即为列表空白处
    m.click(x_dim // 2, y_dim // 2, 1, 1)
    for i in range(rowinfo[0], rowinfo[1]):
        for j in range(0, len(data_array[i])):
            if j < 2:
                yield
                operation_with_pause("pk", k.tab_key, num=3)
                yield
                operation_with_pause("pk", k.delete_key, num=10)
                yield
                operation_with_pause("ps", data_array[i][j][1], num=1)
            else:
                yield
                m.click(x_dim // 2, y_dim // 2, 1, 1)
                yield
                operation_with_pause("pks", k.control_key, "A", num=1)
                yield
                operation_with_pause("ps", data_array[i][j][0], num=1)
                yield
                operation_with_pause("pk", k.tab_key, num=1)
                yield
                operation_with_pause("ps", data_array[i][j][1], num=1)
                yield
                operation_with_pause("pk", k.tab_key, num=2)
                yield
                operation_with_pause("pk", k.up_key, num=1)
                yield
                # # 通过字母位置选择付款方式
                operation_with_pause(
                    "pk", k.up_key, num=pay_mode.index(data_array[i][j][2]))
        if pending == 0:
            yield
            operation_with_pause("pk", k.alt_key, num=1)
            yield
            operation_with_pause("pk", "2", num=1)
            yield
            operation_with_pause("pk", "Y", num=1)
            yield
            # 2次确认弹窗
            time.sleep(stoptime / 2)  # 根据不同电脑的处理速度，选择添加一定的延迟，保证程序正常运行
            # B200或者B202程序要求填入数据，没有填入会弹窗，此处回车用于消除弹窗
            yield
            operation_with_pause("pk", k.enter_key, num=1)
            yield
            time.sleep(stoptime / 2)  # 根据不同电脑的处理速度，选择添加一定的延迟，保证程序正常运行
            yield
            operation_with_pause("pk", k.enter_key, num=1)
        else:
            yield
            operation_with_pause("pks", k.alt_key, "2", "A", num=1)
            # 三次确认弹窗
            yield
            time.sleep(stoptime / 2)  # 根据不同电脑的处理速度，选择添加一定的延迟，保证程序正常运行
            # B200或者B202程序要求填入数据，没有填入会弹窗，此处回车用于消除弹窗
            yield
            operation_with_pause("pk", k.enter_key, num=1)
            yield
            time.sleep(stoptime / 2)  # 根据不同电脑的处理速度，选择添加一定的延迟，保证程序正常运行
            yield
            operation_with_pause("pk", k.enter_key, num=1)
            yield
            time.sleep(stoptime / 2)  # 根据不同电脑的处理速度，选择添加一定的延迟，保证程序正常运行
            yield
            operation_with_pause("pk", k.enter_key, num=1)
        # 新建表单
        yield
        operation_with_pause("pks", k.alt_key, "2", "N", num=1)
        yield
        time.sleep(stoptime / 2)
        yield
        check_window(u"富基融通商业连锁门店管理系统")
        yield
        ensure_App_Foreground(u"富基融通商业连锁门店管理系统")
        yield
        m.click(x_dim // 2, y_dim // 2, 1, 1)

    Stop_thread(keyboard_monitor_thread)
    Stop_thread(os_exit_byEnd_thread)
    Stop_thread(os_exit_thread)


def Judge_system(str_info):
    system_info = platform.platform()
    if str_info in system_info:
        return True
    else:
        return False


def DownLoad(dbnum, dbsize, size):
    global download_ProgressValue

    '''''回调函数 
    dbnum: 已经下载的数据块 
    dbsize: 数据块的大小 
    size: 远程文件的大小 
    '''
    percent = 100.0 * dbnum * dbsize / size
    if percent > 100:
        percent = 100

    download_ProgressValue.set(percent)


def Check_System_Info(screen_width, screen_height):
    system_info = platform.platform()
    if "Windows-7" in system_info or "Windows-10" in system_info:

        return {"geometry": '352x77+' + '%s+%s' % (screen_width, screen_height),
                "maxsize-x": 352,
                "maxsize-y": 101,
                "textwidth": 50,
                "buttonwidth": 43,
                "height": 2,
                "timedefalut": "1-0",
                "rowdefalut": 2
                }
    else:
        return {"geometry": '363x67+' + '%s+%s' % (screen_width, screen_height),
                "maxsize-x": 363,
                "maxsize-y": 99,
                "textwidth": 50,
                "buttonwidth": 45,
                "height": 3,
                "timedefalut": "2-0.1",
                "rowdefalut": 2
                }


def Refresh_Status_label(info):
    global Status_label
    Status_label.set(info)


def loadview():
    global hllDll, VK_CAPITAL, Status_label, ProgressValue, download_ProgressValue
    root = tkinter.Tk()
    root.title('费用录入工具-version:%s' % Version)
    ico = os.getcwd() + r'\ae.ico'
    root.iconbitmap(ico)
    #root.attributes("-alpha", 0.1)
    screen_width = root.winfo_screenwidth() // 2 - 187
    screen_height = root.winfo_screenheight() // 2 - 260

    windows_params = Check_System_Info(
        root.winfo_screenwidth() // 2 - 187, root.winfo_screenheight() // 2 - 260)
    root.geometry(windows_params["geometry"])
    root.maxsize(windows_params["maxsize-x"], windows_params["maxsize-y"])
    root.minsize(windows_params["maxsize-x"], windows_params["maxsize-y"])
    textwidth = windows_params["textwidth"]
    buttonwidth = windows_params["buttonwidth"]
    height = windows_params["height"]
    timedefalut = windows_params["timedefalut"]
    rowdefalut = windows_params["rowdefalut"]

    # 支付二维码显示
    pay_windows = Toplevel()
    pay_windows.title("购买方式")
    pay_windows.iconbitmap(ico)
    path = os.getcwd() + r'\QR_Code.png'
    tkimg = ImageTk.PhotoImage(file=path)
    topLabel = Label(pay_windows, image=tkimg)
    topLabel.pack()
    pay_windows.withdraw()

    # 更新下载安装进度 #
    download_windows = Toplevel()
    download_windows.title("进度...")
    download_windows.iconbitmap(ico)
    download_ProgressValue = DoubleVar()
    download_ProgressValue.set(0.0)
    ttk.Progressbar(download_windows, orient="horizontal",
                    length=352,
                    mode="determinate",
                    variable=download_ProgressValue).grid(column=1,
                                                          row=1,
                                                          sticky=W,
                                                          columnspan=1)
    download_windows.withdraw()

    ######################################################

    v1 = StringVar()
    v1.set("等待导入数据......")
    l1 = Label(root, text="费用数据:", justify=LEFT).grid(
        column=1, row=1, sticky=W)
    textbox1 = Entry(root, font='微软雅黑 -11', bg='darkgray', width=textwidth, state='readonly', textvariable=v1, justify=LEFT).grid(
        column=2, row=1, sticky=N + S + E + W, columnspan=6)
    button1 = Button(root, text="✚", width=6, height=height, command=lambda: get_path(
        v1)).grid(column=7, row=1, sticky=W, rowspan=2)

    v2 = StringVar()
    v2.set(timedefalut)
    l2 = Label(root, text="延迟秒速:", justify=LEFT).grid(
        column=1, row=2, sticky=W)
    textbox2 = Entry(root, font='微软雅黑 -13', width=5, textvariable=v2, justify=RIGHT).grid(
        column=2, row=2, sticky=W)
    l3 = Label(root, text=" 秒 |", justify=LEFT).grid(
        column=3, row=2, sticky=W)

    v3 = StringVar()
    v3.set(rowdefalut)
    l4 = Label(root, text="起始行:", justify=LEFT).grid(
        column=4, row=2, sticky=E)
    textbox3 = Entry(root, font='微软雅黑 -13', width=4, textvariable=v3, justify=RIGHT).grid(
        column=5, row=2, sticky=W)

    chVarDis = IntVar()   # 用来获取复选框是否被勾选，通过chVarDis.get()来获取其的状态,其状态值为int类型 勾选为1  未勾选为0
    del_check = Checkbutton(root, text="直接审核", font='微软雅黑 -11',
                            height=1, variable=chVarDis, state='normal')
    del_check.deselect()
    del_check.grid(column=6, row=2, sticky=N)

    Button(root, text="开始录入", font='微软雅黑 -13 bold', width=buttonwidth, command=lambda: start_input(v2.get(), pay_windows, v3.get(), chVarDis.get())).grid(
        column=1, row=5, sticky=W, columnspan=7)

    Status_label = StringVar()
    Status_label.set("检查注册信息")
    l5 = Label(root, font='微软雅黑 -9',
               bg='lightgray',
               textvariable=Status_label,
               justify=LEFT).grid(column=1,
                                  row=6,
                                  sticky=N + S + E + W,
                                  columnspan=7)

    Button(root, text="检查更新", font='微软雅黑 -8', width=6, command=lambda:Add_Thread(cku.check_update("130.130.200.30",
                                        Software_Name, Version, download_windows, DownLoad))).grid(
        column=7, row=6, sticky=N, columnspan=1)


    Add_thread(lambda: Check_registration_Status_label(
        "http://130.130.200.49", "input-registrationcode.ini", b"0000000000000000"))

    root.mainloop()


def Check_registration_Status_label(ip, filename, keyvalue):
    global isRegistered, UserName, Company, Department
    Registration = ckr.registration_check(ip, filename, keyvalue)
    if Registration[0]:
        isRegistered = True
        UserName = Registration[1]["UserName"]
        Company = Registration[1]["Company"]
        Department = Registration[1]["Department"]
        Refresh_Status_label("...已激活...")
    else:
        isRegistered = False
        Refresh_Status_label("...未激活...")

if __name__ == '__main__':
    loadview()
